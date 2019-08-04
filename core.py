import csv, datetime, decimal
from time import sleep
from openpyxl import load_workbook
from collections import Counter


class personal_budgeter(object):

    def __init__(self):

        self.expenses, self.income = [], []

        self.wb_file = 'exts/personal_budget_template.xlsx'

        expense_counter = Counter({
            'Bitcoin': 0,
            'Amazon': 0,
            'Miscellaneous': 0
        })

        self.expense_headers = {
            'Amazon': ["Amazon", "AMZN Mktp US", 'Amazon.com'],
            'Bitcoin': ['purchase of BTC']
        }

        self.budget_workbook = load_workbook(self.wb_file)
        self.personal_budget_sheet = self.budget_workbook.active
        self.rows = [row for row in self.personal_budget_sheet]
        self.date_row = [date for date in self.rows[1] if isinstance(date.value, datetime.datetime)]
        self.income_rows = self.rows[2][1]
        self.income_rows = self.column_select_to(self.income_rows, 'Total Income')
        self.expense_rows = self.rows[self.income_rows[-1].row+1][1]
        self.expense_rows = self.column_select_to(self.expense_rows, 'Total Expenses', 20)
        self.month_expenses = {i.value.strftime('%b-%Y'): expense_counter.copy() for i in self.date_row}
        self.create_headers(self.expense_rows[0], list(expense_counter.keys()))


    def create_headers(self, m_header, h_titles):
        
        orgin_coordinates = (m_header.column, m_header.row)

        for title in h_titles:
            #import pdb; pdb.set_trace(header='h_title break')
            m_header.row += 1;
            print(f'Making edit to cell: {m_header.coordinate}')
            self.personal_budget_sheet[m_header.coordinate] = title

        self.budget_workbook.save(self.wb_file)
        m_header.column, m_header.row = orgin_coordinates
        #import pdb; pdb.set_trace()

    def append_expense(self, transaction, key):

        #import pdb; pdb.set_trace(header='Append Expense Break')

        # Add specific expense for month
        self.month_expenses[datetime.datetime.strptime(\
            transaction['Date'].strip(' CDT CST'), \
            '%Y-%m-%d %H:%M:%S').strftime('%b-%Y'\
        )][key] += float(transaction['Amount'].strip('-$'))
        #print(f"Adding {float(transaction['Amount'].strip('-$'))}")

        #import pdb; pdb.set_trace()

    def load_square_report(self, rfile):

        with open(rfile) as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                if row['Status'] == 'CARD CHARGED': self.expenses.append(row)
                elif row['Status'] == 'PAYMENT DEPOSITED': self.income.append(row)

        temp = []
        for transaction in self.expenses:
            transaction_date = datetime.datetime.strptime(transaction['Date'].strip(' CDT CST'), '%Y-%m-%d %H:%M:%S')
            if (self.date_row[0].value <= transaction_date <= self.date_row[-1].value): temp.append(transaction)

        self.expenses = temp; del temp

        self.calculate_expense(self.expenses)

    def calculate_expense(self, exp):

        expense_total = sum(float(amount['Amount'].strip('-$')) for amount in exp if amount['Currency'] == 'USD')

        for transaction in exp:

            misc = True
            for key, value in self.expense_headers.items():

                if any(search_tag in transaction['Notes'] for search_tag in value) and (transaction['Currency'] == 'USD'):
                    self.append_expense(transaction, key); misc = False

            if misc and (transaction['Currency'] == 'USD'): self.append_expense(transaction, 'Miscellaneous')

    def column_select_to(self, cell, strtarget, scan_range=10):

        scanned_cells = [self.rows[i][cell.col_idx - 1] for i in range(scan_range)]
        cell_vals = [cell.value for cell in scanned_cells]
        scanned_cells = scanned_cells[(cell.row-1):(cell_vals.index(strtarget)+1)]

        return(scanned_cells)

        #import pdb; pdb.set_trace(header="Column Break")

    def generate_budget_report(self):
        
        for i, month in enumerate(self.date_row):

            #import pdb; pdb.set_trace(header='Before dictionary break')
            current_month = month.value.strftime("%b-%Y")
            #print(f'If {current_month} == {list(self.month_expenses.keys())[i]}')
            if current_month == list(self.month_expenses.keys())[i]:

                expense_values = [row.value for row in self.expense_rows if row.value != None][:-1]
                import pdb; pdb.set_trace()

                for n, header in enumerate(expense_values):
                    target_cell = self.personal_budget_sheet.cell(row=self.expense_rows[(n+1)].row, column=month.column)
                    target_cell.value = list(self.month_expenses.values())[i][header]
                    print(list(self.month_expenses.values())[i][header], target_cell)
                    self.budget_workbook.save(self.wb_file)

                #import pdb; pdb.set_trace(header='Expenses Break')

if __name__ == '__main__':
    app = personal_budgeter()
    app.load_square_report('exts/square_cash_report.csv')
    app.generate_budget_report()

    import pdb; pdb.set_trace()